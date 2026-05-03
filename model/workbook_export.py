"""Excel review-workbook exporter.

Generates `outputs/workbooks/ai_economy_model_review.xlsx` from the
existing CSV outputs (no DuckDB dependency — workbook reads CSVs
directly so it works with or without a database build).

11 sheets per the upstream review-layer scope:
  1. README                     — how to read the workbook
  2. Model Flow                 — component flow + downstream consumers
  3. Scenario Matrix            — 16 combined scenarios × milestone metrics
  4. Historical Baseline        — historical_trend_estimates summary
  5. Supply Capacity            — supply_scenario_summary + capex
  6. Allocation Buckets         — allocation_compute_by_bucket (all rows)
  7. Largest Frontier Run       — allocation_largest_frontier_run (all rows)
  8. Phase 4 Handoff            — slow / base / fast envelope, side-by-side
  9. Assumptions                — share assumptions per year + supply summary
 10. Sources & Confidence       — flattened YAML metadata
 11. Output Inventory           — every artifact + path + description

Workbook is a generated review artifact; source of truth remains
YAML / Python / CSV / Markdown.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from model.review_database import (
    ALLOCATION_SCENARIO_ID,
    SUPPLY_SCENARIO_ID,
    _flatten_allocation_assumptions,
    _flatten_supply_assumptions,
)
from model.runtime import (
    ASSUMPTIONS_DIR,
    CHARTS_DIR,
    OUTPUTS_DIR,
    PROCESSED_DIR,
    TABLES_DIR,
)

WORKBOOK_DIR = OUTPUTS_DIR / "workbooks"
WORKBOOK_PATH = WORKBOOK_DIR / "ai_economy_model_review.xlsx"

# --- styling primitives -------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")

CONFIDENCE_FILLS = {
    "high": PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "low": PatternFill("solid", fgColor="FFC7CE"),
}

NUM_FORMATS = {
    "scientific": "0.00E+00",
    "percent": "0.0%",
    "dollars": "$0.00E+00",
    "ratio": "0.00",
    "year": "0",
    "count": "#,##0",
}


def _style_header_row(ws: Worksheet, n_cols: int, row: int = 1) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws: Worksheet, min_width: int = 12, max_width: int = 40) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[col_letter].width = max(
            min_width, min(max_width, max_len + 2)
        )


def _apply_format_to_column(ws: Worksheet, col_idx: int, num_format: str) -> None:
    """Apply a number format to all data cells in a 1-indexed column."""
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_idx).number_format = num_format


def _write_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    *,
    start_row: int = 1,
    freeze: bool = True,
    autofilter: bool = True,
) -> int:
    """Write df to worksheet starting at `start_row`. Returns the next free row."""
    for row_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=start_row
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _style_header_row(ws, len(df.columns), row=start_row)
    if freeze and start_row == 1:
        ws.freeze_panes = "A2"
    if autofilter and len(df) > 0:
        ws.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"
        )
    return start_row + len(df) + 1


# --- per-sheet builders -------------------------------------------------


def _sheet_readme(wb: Workbook, manifest: dict) -> None:
    ws = wb.active
    ws.title = "README"
    lines = [
        ("ai-economy-timelines — Model Review Workbook", "title"),
        ("", None),
        (f"Generated: {manifest['generated_at']}", None),
        (f"Repository commit: {manifest['git_commit']}", None),
        (f"Schema version: {manifest['schema_version']}", None),
        ("", None),
        ("What this workbook is", "section"),
        ("A generated review artifact summarizing the model's outputs across", None),
        ("11 sheets. Built from the CSVs in outputs/tables/ and the assumption", None),
        ("YAMLs under data/assumptions/. Reproducible end-to-end via:", None),
        ("    uv run historical && uv run supply && uv run allocation", None),
        ("    uv run database  &&  uv run workbook", None),
        ("", None),
        ("Sheet index", "section"),
        ("  1.  README                  — this sheet", None),
        ("  2.  Model Flow              — component flow + downstream consumers", None),
        ("  3.  Scenario Matrix         — 16 combined scenarios × milestone metrics", None),
        ("  4.  Historical Baseline     — frontier-run trend fits", None),
        ("  5.  Supply Capacity         — supply scenario summary + capex", None),
        ("  6.  Allocation Buckets      — bucket-level compute by combined scenario", None),
        ("  7.  Largest Frontier Run    — headline largest_frontier_run_flop", None),
        ("  8.  Phase 4 Handoff         — slow / base / fast envelope, side-by-side", None),
        ("  9.  Assumptions             — share assumptions + supply summary", None),
        (" 10.  Sources & Confidence    — flattened YAML metadata", None),
        (" 11.  Output Inventory        — every artifact + path + description", None),
        ("", None),
        ("Source of truth", "section"),
        ("This workbook is a generated review artifact. Source of truth", None),
        ("remains:", None),
        ("  - YAML assumptions          (data/assumptions/*.yaml)", None),
        ("  - Python pipeline modules   (pipelines/*, model/*)", None),
        ("  - CSV / Parquet outputs     (outputs/tables/, data/processed/)", None),
        ("  - Markdown findings         (docs/*.md)", None),
        ("", None),
        ("Do NOT edit this workbook by hand. Edit the upstream sources and", None),
        ("regenerate via `uv run workbook`.", None),
        ("", None),
        ("What NOT to infer from this workbook", "section"),
        ("  - largest_frontier_run_flop is a single-training-run quantity.", None),
        ("    It is NOT total annual usable AI compute. The two are not", None),
        ("    interchangeable as forecasts.", None),
        ("  - All FLOP figures are RAW. They have not been adjusted for", None),
        ("    algorithmic-efficiency gains. The effective-compute layer", None),
        ("    (next, not yet built) will produce algorithm-adjusted figures.", None),
        ("  - Allocation parameters are flagged confidence: medium with", None),
        ("    source: scope_defaults. They are not yet sourced from", None),
        ("    hyperscaler 10-Ks or lab disclosures.", None),
        ("", None),
        ("More context", "section"),
        ("  - docs/executive_summary.md  — plain-English summary", None),
        ("  - docs/model_map.md          — architecture + data flow", None),
        ("  - docs/output_guide.md       — per-file interpretation", None),
        ("  - docs/review_workbook_guide.md — how to use this workbook", None),
    ]
    for r, (text, kind) in enumerate(lines, start=1):
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            cell.font = Font(bold=True, size=16, color="1F4E78")
        elif kind == "section":
            cell.font = SECTION_FONT
    ws.column_dimensions["A"].width = 90


def _sheet_model_flow(wb: Workbook) -> None:
    ws = wb.create_sheet("Model Flow")
    rows = [
        ["Component", "Status", "Inputs", "Transformation", "Outputs", "Downstream consumer"],
        ["Historical baseline", "BUILT",
         "Epoch Notable AI Models CSV",
         "Frontier filters + log-linear fits",
         "historical_trend_estimates.csv",
         "Calibration target (allocation layer)"],
        ["Supply capacity model", "BUILT",
         "supply_input_assumptions.yaml + 4 scenarios",
         "Shipments → stock → 4 limits → utilization derating",
         "supply_fundamental_inputs_by_year.csv",
         "Allocation layer"],
        ["Allocation layer", "BUILT",
         "supply outputs + allocation_input_assumptions.yaml",
         "16 combined scenarios; bucket allocation; training decomp; largest-run estimate",
         "allocation_largest_frontier_run.csv",
         "Effective-compute layer"],
        ["Effective compute", "NEXT",
         "largest_frontier_run_flop_by_year + algorithmic-efficiency assumptions",
         "Adjust raw FLOP for algorithmic efficiency gains",
         "(planned) effective_compute_by_year.csv",
         "Capability mapping"],
        ["Capability mapping", "FUTURE", "—", "—", "—", "—"],
        ["Projection engine", "FUTURE", "—", "—", "—", "—"],
        ["Economy feedback", "FUTURE", "—", "—", "—", "—"],
    ]
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _style_header_row(ws, len(rows[0]))
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 38

    # Status fill: green/orange/grey
    for r in range(2, len(rows) + 1):
        status = ws.cell(row=r, column=2).value
        if status == "BUILT":
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="C6EFCE")
        elif status == "NEXT":
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFEB9C")
        elif status == "FUTURE":
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="DDDDDD")


def _sheet_scenario_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Scenario Matrix")
    df = pd.read_csv(TABLES_DIR / "allocation_scenario_summary.csv")
    # Add combined_scenario_id
    df["combined_scenario_id"] = (
        df["supply_scenario"].map(SUPPLY_SCENARIO_ID).fillna(df["supply_scenario"])
        + "__"
        + df["allocation_scenario"].map(ALLOCATION_SCENARIO_ID).fillna(df["allocation_scenario"])
    )
    # Pull binding constraint from supply outputs
    supply = pd.read_csv(TABLES_DIR / "supply_fundamental_inputs_by_year.csv")
    binding_2030 = supply[supply["year"] == 2030].set_index("scenario")["binding_constraint"]
    binding_2040 = supply[supply["year"] == 2040].set_index("scenario")["binding_constraint"]
    df["binding_constraint_2030"] = df["supply_scenario"].map(binding_2030)
    df["binding_constraint_2040"] = df["supply_scenario"].map(binding_2040)
    df = df[[
        "combined_scenario_id", "combined_scenario", "supply_scenario", "allocation_scenario",
        "usable_compute_2030", "usable_compute_2040",
        "largest_run_2030", "largest_run_2040", "largest_run_cagr_2024_2040",
        "frontier_run_share_2030", "frontier_run_share_2040",
        "binding_constraint_2030", "binding_constraint_2040",
    ]].sort_values("largest_run_2040", ascending=False).reset_index(drop=True)

    _write_dataframe(ws, df)

    # Number formats
    cols = list(df.columns)
    sci_cols = ["usable_compute_2030", "usable_compute_2040",
                "largest_run_2030", "largest_run_2040"]
    pct_cols = ["largest_run_cagr_2024_2040",
                "frontier_run_share_2030", "frontier_run_share_2040"]
    for col_name in sci_cols:
        _apply_format_to_column(ws, cols.index(col_name) + 1, NUM_FORMATS["scientific"])
    for col_name in pct_cols:
        _apply_format_to_column(ws, cols.index(col_name) + 1, NUM_FORMATS["percent"])

    # Color scale on largest_run_2040
    largest_col = get_column_letter(cols.index("largest_run_2040") + 1)
    ws.conditional_formatting.add(
        f"{largest_col}2:{largest_col}{len(df) + 1}",
        ColorScaleRule(
            start_type="min", start_color="FFC7CE",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE",
        ),
    )
    _autosize_columns(ws)


def _sheet_historical_baseline(wb: Workbook) -> None:
    ws = wb.create_sheet("Historical Baseline")
    df = pd.read_csv(TABLES_DIR / "historical_trend_estimates.csv")
    _write_dataframe(ws, df)
    cols = list(df.columns)
    if "annual_growth_multiplier" in cols:
        _apply_format_to_column(
            ws, cols.index("annual_growth_multiplier") + 1, NUM_FORMATS["ratio"]
        )
    if "r_squared" in cols:
        _apply_format_to_column(ws, cols.index("r_squared") + 1, NUM_FORMATS["ratio"])
    if "doubling_time_years" in cols:
        _apply_format_to_column(ws, cols.index("doubling_time_years") + 1, NUM_FORMATS["ratio"])
    _autosize_columns(ws)


def _sheet_supply_capacity(wb: Workbook) -> None:
    ws = wb.create_sheet("Supply Capacity")
    summary = pd.read_csv(TABLES_DIR / "supply_scenario_summary.csv")
    capex = pd.read_csv(TABLES_DIR / "supply_capex_requirements.csv")

    # Section 1: scenario summary
    ws.cell(row=1, column=1, value="Supply scenario summary").font = SECTION_FONT
    next_row = _write_dataframe(ws, summary, start_row=2, freeze=False, autofilter=False)

    # Section 2: capex requirements
    ws.cell(row=next_row + 1, column=1, value="Capex required vs available").font = SECTION_FONT
    _write_dataframe(ws, capex, start_row=next_row + 2, freeze=False, autofilter=False)

    cols = list(summary.columns)
    for sci_col in ["usable_compute_flop_year", "ai_power_capacity_mw",
                    "ai_infrastructure_capex_usd", "available_stock_h100e"]:
        if sci_col in cols:
            for r in range(3, 3 + len(summary)):
                ws.cell(row=r, column=cols.index(sci_col) + 1).number_format = NUM_FORMATS["scientific"]
    _autosize_columns(ws)


def _sheet_allocation_buckets(wb: Workbook) -> None:
    ws = wb.create_sheet("Allocation Buckets")
    df = pd.read_csv(TABLES_DIR / "allocation_compute_by_bucket.csv")
    df["combined_scenario_id"] = (
        df["supply_scenario"].map(SUPPLY_SCENARIO_ID).fillna(df["supply_scenario"])
        + "__"
        + df["allocation_scenario"].map(ALLOCATION_SCENARIO_ID).fillna(df["allocation_scenario"])
    )
    leading = ["combined_scenario_id", "year", "supply_scenario", "allocation_scenario"]
    rest = [c for c in df.columns if c not in leading + ["combined_scenario"]]
    df = df[leading + rest]
    _write_dataframe(ws, df)
    bucket_cols = [c for c in df.columns if c.endswith("_compute_flop_year")]
    cols = list(df.columns)
    for c in bucket_cols + ["usable_compute_flop_year",
                            "frontier_lab_training_compute_flop_year",
                            "largest_frontier_run_flop"]:
        if c in cols:
            _apply_format_to_column(ws, cols.index(c) + 1, NUM_FORMATS["scientific"])
    for c in [c for c in df.columns if c.endswith("_share")] + [
        "frontier_lab_training_share", "largest_run_concentration",
        "cluster_contiguity_factor", "frontier_run_share_of_total_compute",
    ]:
        if c in cols:
            _apply_format_to_column(ws, cols.index(c) + 1, NUM_FORMATS["percent"])
    _autosize_columns(ws)


def _sheet_largest_frontier_run(wb: Workbook) -> None:
    ws = wb.create_sheet("Largest Frontier Run")
    df = pd.read_csv(TABLES_DIR / "allocation_largest_frontier_run.csv")
    df["combined_scenario_id"] = (
        df["supply_scenario"].map(SUPPLY_SCENARIO_ID).fillna(df["supply_scenario"])
        + "__"
        + df["allocation_scenario"].map(ALLOCATION_SCENARIO_ID).fillna(df["allocation_scenario"])
    )
    df = df[["combined_scenario_id", "year", "combined_scenario",
             "supply_scenario", "allocation_scenario",
             "largest_frontier_run_flop", "frontier_run_share_of_total_compute",
             "training_compute_flop_year",
             "frontier_lab_training_compute_flop_year"]]
    _write_dataframe(ws, df)
    cols = list(df.columns)
    for c in ["largest_frontier_run_flop", "training_compute_flop_year",
              "frontier_lab_training_compute_flop_year"]:
        _apply_format_to_column(ws, cols.index(c) + 1, NUM_FORMATS["scientific"])
    _apply_format_to_column(
        ws, cols.index("frontier_run_share_of_total_compute") + 1, NUM_FORMATS["percent"]
    )

    # Highlight the slow / base / fast rows
    envelope_ids = {
        "chip_bottleneck__inference_heavy": "FFC7CE",  # slow
        "base__base": "FFEB9C",                         # base
        "capex_rich__training_race": "C6EFCE",          # fast
    }
    id_col_idx = cols.index("combined_scenario_id") + 1
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=id_col_idx).value
        if sid in envelope_ids:
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(
                    "solid", fgColor=envelope_ids[sid]
                )

    _autosize_columns(ws)


def _sheet_phase4_handoff(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase 4 Handoff")
    df = pd.read_csv(TABLES_DIR / "allocation_compute_by_bucket.csv")
    df["combined_scenario_id"] = (
        df["supply_scenario"].map(SUPPLY_SCENARIO_ID).fillna(df["supply_scenario"])
        + "__"
        + df["allocation_scenario"].map(ALLOCATION_SCENARIO_ID).fillna(df["allocation_scenario"])
    )

    # Pivot slow/base/fast side-by-side
    handoff_ids = {
        "chip_bottleneck__inference_heavy": "slow",
        "base__base": "base",
        "capex_rich__training_race": "fast",
    }
    sub = df[df["combined_scenario_id"].isin(handoff_ids.keys())].copy()
    sub["envelope"] = sub["combined_scenario_id"].map(handoff_ids)

    # Wide format on largest_frontier_run_flop
    largest = (
        sub.pivot(index="year", columns="envelope", values="largest_frontier_run_flop")
        [["slow", "base", "fast"]]
        .rename(columns={
            "slow": "slow_largest_frontier_run_flop",
            "base": "base_largest_frontier_run_flop",
            "fast": "fast_largest_frontier_run_flop",
        })
    )
    base_only = sub[sub["envelope"] == "base"].set_index("year")[[
        "training_compute_flop_year",
        "ai_rnd_experiment_compute_flop_year",
        "post_training_compute_flop_year",
        "inference_compute_flop_year",
    ]].rename(columns={
        "training_compute_flop_year": "base_training_compute_flop_year",
        "ai_rnd_experiment_compute_flop_year": "base_ai_rnd_experiment_compute_flop_year",
        "post_training_compute_flop_year": "base_post_training_compute_flop_year",
        "inference_compute_flop_year": "base_inference_compute_flop_year",
    })
    handoff = largest.join(base_only).reset_index()

    # Header note
    ws.cell(row=1, column=1, value=(
        "Slow / Base / Fast envelope handoff into the effective-compute layer. "
        "Slow = chip_bottleneck × inference_heavy; "
        "Base = base × base; "
        "Fast = capex_rich × training_race. "
        "Bucket totals shown for the Base case only."
    )).font = Font(italic=True, color="666666")
    ws.row_dimensions[1].height = 32
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    _write_dataframe(ws, handoff, start_row=3, freeze=False, autofilter=False)
    ws.freeze_panes = "B4"  # year column locked
    cols = list(handoff.columns)
    for c in cols:
        if c == "year":
            continue
        _apply_format_to_column(ws, cols.index(c) + 1, NUM_FORMATS["scientific"])
        # Note: _apply_format_to_column starts at row 2 by default; adjust:
    # Re-apply with correct row offset (header is row 3, data starts row 4)
    for c in cols:
        if c == "year":
            for r in range(4, 4 + len(handoff)):
                ws.cell(row=r, column=cols.index(c) + 1).number_format = NUM_FORMATS["year"]
        else:
            for r in range(4, 4 + len(handoff)):
                ws.cell(row=r, column=cols.index(c) + 1).number_format = NUM_FORMATS["scientific"]

    # Color-code the slow / base / fast columns
    color_map = {
        "slow_largest_frontier_run_flop": "FFC7CE",
        "base_largest_frontier_run_flop": "FFEB9C",
        "fast_largest_frontier_run_flop": "C6EFCE",
    }
    for col_name, color in color_map.items():
        cidx = cols.index(col_name) + 1
        ws.cell(row=3, column=cidx).fill = PatternFill("solid", fgColor=color)
    _autosize_columns(ws)


def _sheet_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Assumptions")
    alloc = pd.read_csv(TABLES_DIR / "allocation_share_assumptions_by_year.csv")
    ws.cell(row=1, column=1, value="Allocation share assumptions by year (interpolated)").font = SECTION_FONT
    nxt = _write_dataframe(ws, alloc, start_row=2, freeze=False, autofilter=False)
    cols = list(alloc.columns)
    for c in cols:
        if c.endswith("_share") or c in (
            "frontier_lab_training_share",
            "largest_run_concentration",
            "cluster_contiguity_factor",
        ):
            for r in range(3, 3 + len(alloc)):
                ws.cell(row=r, column=cols.index(c) + 1).number_format = NUM_FORMATS["percent"]

    # Supply assumption summary (high-level — number of milestones per param)
    ws.cell(row=nxt + 1, column=1, value="Supply assumptions summary").font = SECTION_FONT
    rows = _flatten_supply_assumptions(ASSUMPTIONS_DIR / "supply_input_assumptions.yaml")
    if rows:
        sdf = pd.DataFrame(rows)[["input", "scenario", "source", "confidence", "unit", "notes"]]
        _write_dataframe(ws, sdf, start_row=nxt + 2, freeze=False, autofilter=False)
        # Confidence colouring
        conf_col = list(sdf.columns).index("confidence") + 1
        for r in range(nxt + 3, nxt + 3 + len(sdf)):
            v = ws.cell(row=r, column=conf_col).value
            if v in CONFIDENCE_FILLS:
                ws.cell(row=r, column=conf_col).fill = CONFIDENCE_FILLS[v]
    _autosize_columns(ws)


def _sheet_sources_and_confidence(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources & Confidence")
    rows = _flatten_supply_assumptions(ASSUMPTIONS_DIR / "supply_input_assumptions.yaml")
    rows.extend(_flatten_allocation_assumptions(ASSUMPTIONS_DIR / "allocation_input_assumptions.yaml"))
    df = pd.DataFrame(rows)[
        ["input", "component", "scenario", "source", "source_type", "confidence", "unit", "used_for", "notes"]
    ]
    _write_dataframe(ws, df)
    conf_col = list(df.columns).index("confidence") + 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=conf_col).value
        if v in CONFIDENCE_FILLS:
            ws.cell(row=r, column=conf_col).fill = CONFIDENCE_FILLS[v]
    _autosize_columns(ws)


def _sheet_output_inventory(wb: Workbook) -> None:
    ws = wb.create_sheet("Output Inventory")
    rows = []
    # Tables
    for path in sorted(TABLES_DIR.glob("*.csv")):
        component = (
            "historical" if path.name.startswith("historical_")
            else "supply" if path.name.startswith("supply_")
            else "allocation" if path.name.startswith("allocation_")
            else "review"
        )
        rows.append({
            "artifact": path.name,
            "path": f"outputs/tables/{path.name}",
            "component": component,
            "type": "table",
            "description": f"Output table from the {component} pipeline",
            "generated_by": f"pipelines/{component}.py" if component != "review" else "pipelines/build_review_database.py",
            "downstream_use": "review workbook + DuckDB",
        })
    # Charts
    for path in sorted(CHARTS_DIR.glob("*.png")):
        component = (
            "historical" if path.name.startswith("historical_")
            else "supply" if path.name.startswith("supply_")
            else "allocation" if path.name.startswith("allocation_")
            else "review"
        )
        rows.append({
            "artifact": path.name,
            "path": f"outputs/charts/{path.name}",
            "component": component,
            "type": "chart",
            "description": f"PNG from the {component} pipeline",
            "generated_by": f"pipelines/{component}.py",
            "downstream_use": "docs + review",
        })
    # Processed datasets
    for path in sorted(PROCESSED_DIR.glob("*")):
        rows.append({
            "artifact": path.name,
            "path": f"data/processed/{path.name}",
            "component": (
                "historical" if "historical" in path.name
                else "supply" if "supply" in path.name
                else "allocation" if "allocation" in path.name
                else "unknown"
            ),
            "type": "processed dataset",
            "description": "Cleaned dataset used downstream",
            "generated_by": "(corresponding pipeline)",
            "downstream_use": "downstream pipelines + review",
        })
    df = pd.DataFrame(rows)
    _write_dataframe(ws, df)
    _autosize_columns(ws)


# --- top-level orchestrator ---------------------------------------------


def export_workbook(out: Path = WORKBOOK_PATH) -> dict:
    """Build the 11-sheet Excel review workbook. Returns a small manifest dict."""
    out.parent.mkdir(parents=True, exist_ok=True)

    manifest = {
        "schema_version": "1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "git_commit": _git_commit(),
        "workbook_path": str(out.relative_to(out.parent.parent.parent)),
    }

    wb = Workbook()
    _sheet_readme(wb, manifest)
    _sheet_model_flow(wb)
    _sheet_scenario_matrix(wb)
    _sheet_historical_baseline(wb)
    _sheet_supply_capacity(wb)
    _sheet_allocation_buckets(wb)
    _sheet_largest_frontier_run(wb)
    _sheet_phase4_handoff(wb)
    _sheet_assumptions(wb)
    _sheet_sources_and_confidence(wb)
    _sheet_output_inventory(wb)
    wb.save(out)
    return manifest


def _git_commit() -> str:
    import subprocess
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
            cwd=OUTPUTS_DIR.parent,
        ).decode().strip()
    except Exception:
        return "unknown"
